import os
import pandas as pd
import shutil
from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, F
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import JsonResponse
from datetime import timedelta
from django.conf import settings
from django.http import FileResponse

from .models import FollowUp, MasterCUP
from .forms import FollowUpForm, UploadFileForm
from .services import (
    importar_archivo_masivo, 
    compute_request_status_from_db, 
    compute_opportunity_by_procedure,
    compute_barriers,
    compute_institutional_metrics_db
)

@login_required
def sembrar_cups(request):
    """
    Ingeniería Inversa: Recorre los seguimientos existentes, extrae los CUPS
    y los guarda en el Maestro para que sean clasificados.
    """
    if not request.user.is_superuser:
        return redirect('followup_dashboard')
        
    # Traemos todos los pares únicos (CUPS, Procedimiento Original)
    existentes = FollowUp.objects.exclude(cups__isnull=True).exclude(cups='').values('cups', 'tipo_procedimiento').distinct()
    
    creados = 0
    for item in existentes:
        codigo = item['cups'].strip().upper()
        desc = item['tipo_procedimiento']
        
        # Intentamos adivinar el grupo basado en lo que ya dice el sistema
        grupo_inicial = 'PENDIENTE'
        if 'CONSULTA' in desc: grupo_inicial = 'CONSULTA'
        elif 'LABORATORIO' in desc: grupo_inicial = 'LABORATORIO'
        elif 'IMAGEN' in desc: grupo_inicial = 'IMAGEN'
        elif 'QUIMIO' in desc: grupo_inicial = 'QUIMIOTERAPIA'
        elif 'RADIO' in desc: grupo_inicial = 'RADIOTERAPIA'
        elif 'CIRUGIA' in desc: grupo_inicial = 'CIRUGIA'

        obj, created = MasterCUP.objects.get_or_create(
            codigo=codigo,
            defaults={
                'descripcion': desc,
                'grupo': grupo_inicial
            }
        )
        if created: creados += 1

    messages.success(request, f"🌱 Cosecha finalizada. Se aprendieron {creados} códigos nuevos.")
    return redirect('followup_dashboard')

# --- 1. TABLERO PRINCIPAL (DASHBOARD) ---
@login_required
def followup_dashboard(request):
    """
    Vista principal con FILTROS PERSISTENTES (Memoria de Sesión).
    """
    # 1. LÓGICA DE MEMORIA (PERSISTENCIA)
    # Recuperamos filtros guardados o usamos diccionario vacío
    filtros_guardados = request.session.get('filtros_dashboard', {})
    
    # Determinamos qué filtros usar:
    if request.GET:
        # A. Si el usuario está enviando datos nuevos (hizo clic en Filtrar o Paginación)
        if 'limpiar_filtros' in request.GET:
            # Caso especial: Botón Limpiar
            request.session['filtros_dashboard'] = {}
            return redirect('followup_dashboard')
        else:
            # Caso normal: Guardamos la nueva selección en sesión
            # Convertimos QueryDict a dict normal para poder guardarlo
            filtros_guardados = {k: v for k, v in request.GET.items() if v}
            request.session['filtros_dashboard'] = filtros_guardados
    else:
        # B. Si no hay GET (Viene de "Atrás" o menú), usamos la memoria
        pass # filtros_guardados ya tiene lo de la sesión

    # Usamos 'params' en lugar de 'request.GET' de aquí en adelante
    params = filtros_guardados

    # 2. Queryset Base
    queryset = FollowUp.objects.select_related('patient').all().order_by('-fecha_solicitud_cita', '-id')

    # 3. Captura de Filtros (Desde 'params', no request.GET)
    date_from = params.get('date_from')
    date_to = params.get('date_to')
    status = params.get('status')
    procedure = params.get('procedure')
    eps = params.get('eps')
    barrier = params.get('barrier')
    agrupador = params.get('agrupador')
    q_search = params.get('q')

    # 4. Aplicación de Filtros
    if date_from: queryset = queryset.filter(fecha_solicitud_cita__gte=date_from)
    if date_to: queryset = queryset.filter(fecha_solicitud_cita__lte=date_to)
    
    if status: queryset = queryset.filter(estado_solicitud__icontains=status)
    if procedure: queryset = queryset.filter(tipo_procedimiento__icontains=procedure)
    
    if eps: queryset = queryset.filter(entidad_aseguradora=eps)
    if barrier: queryset = queryset.filter(barrera=barrier)
    if agrupador: queryset = queryset.filter(agrupador=agrupador)
    
    if q_search:
        queryset = queryset.filter(
            Q(patient__numero_documento__icontains=q_search) |
            Q(patient__nombre_1__icontains=q_search) |          
            Q(patient__apellido_1__icontains=q_search) |       
            Q(observaciones__icontains=q_search) |              
            Q(cups__icontains=q_search)
        )

    # 5. Listas para Dropdowns
    eps_options = FollowUp.objects.exclude(entidad_aseguradora__isnull=True).exclude(entidad_aseguradora='').values_list('entidad_aseguradora', flat=True).distinct().order_by('entidad_aseguradora')
    barrier_options = FollowUp.objects.exclude(barrera__isnull=True).exclude(barrera='').values_list('barrera', flat=True).distinct().order_by('barrera')
    agrupador_options = FollowUp.objects.exclude(agrupador__isnull=True).exclude(agrupador='').values_list('agrupador', flat=True).distinct().order_by('agrupador')

    # 6. Estadísticas (Sobre el queryset ya filtrado)
    grand_total = FollowUp.objects.count()
    total_registros_filtrados = queryset.count()
    
    pct_global = 0
    if grand_total > 0:
        pct_global = round((total_registros_filtrados / grand_total) * 100, 1)

    kpi_status = compute_request_status_from_db(queryset)
    kpi_procedure = compute_opportunity_by_procedure(queryset)
    kpi_barriers = compute_barriers(queryset)

    pct_pendientes = 0
    if total_registros_filtrados > 0:
        pct_pendientes = round((kpi_status['pendientes'] / total_registros_filtrados) * 100, 1)

    realizados = queryset.filter(estado_solicitud__icontains='REALIZADO', fecha_cita__isnull=False, fecha_solicitud_cita__isnull=False)
    dias_list = [r.dias_diff for r in realizados if r.dias_diff is not None and r.dias_diff >= 0]
    promedio = round(sum(dias_list) / len(dias_list), 1) if dias_list else 0

    stats = {
        'total': total_registros_filtrados,
        'pct_global': pct_global,
        'grand_total': grand_total,           
        'completados': kpi_status['completados'],
        'pendientes': kpi_status['pendientes'],
        'porcentaje_completado': kpi_status['porcentaje_completado'],
        'porcentaje_pendientes': pct_pendientes,
        'promedio_dias': promedio
    }

    # 7. Paginación
    per_page = params.get('per_page', 25) # También recordamos cuántas filas le gusta ver
    try: per_page = int(per_page)
    except: per_page = 25

    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get('page') # La página sí debe ser fresca siempre
    page_obj = paginator.get_page(page_number)

    context = {
        'per_page': per_page, 
        'page_obj': page_obj,
        'filtros': params, # <--- ENVIAMOS LOS FILTROS DE MEMORIA AL HTML
        'stats': stats,
        'eps_options': eps_options,       
        'barrier_options': barrier_options,
        'agrupador_options': agrupador_options,
        'estado_procedimiento_labels': kpi_status['labels'],
        'estado_procedimiento_values': kpi_status['values'],
        'oportunidad_procedimiento_labels': kpi_procedure['procedimiento_labels'],
        'oportunidad_procedimiento_values': kpi_procedure['values'],
        'barreras_labels': kpi_barriers['labels'],
        'barreras_values': kpi_barriers['values'],
    }
    return render(request, 'followups.html', context)

# --- 2. IMPORTACIÓN DE DATOS ---
@login_required
def importar_datos(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            fs = FileSystemStorage()
            filename = fs.save(archivo.name, archivo)
            file_path = fs.path(filename)
            try:
                resultado = importar_archivo_masivo(file_path)
                if resultado.get('success'):
                    msg = f"Carga completada. Nuevos: {resultado.get('registros')}, Actualizados: {resultado.get('actualizados')}"
                    messages.success(request, msg)
                else:
                    messages.error(request, f"Error: {resultado.get('error')}")
            except Exception as e:
                messages.error(request, f"Error crítico: {str(e)}")
            finally:
                if os.path.exists(file_path):
                    os.remove(file_path)
            return redirect('followup_dashboard')
    else:
        form = UploadFileForm()
    return render(request, 'cargar_datos.html', {'form': form})

# --- 3. EXPORTACIÓN EXCEL PROFESIONAL ---
@login_required
def exportar_excel(request):
    queryset = FollowUp.objects.select_related('patient').all().order_by('-fecha_solicitud_cita')
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status = request.GET.get('status')
    procedure = request.GET.get('procedure')
    eps = request.GET.get('eps')
    barrier = request.GET.get('barrier')
    q_search = request.GET.get('q')

    if date_from: queryset = queryset.filter(fecha_solicitud_cita__gte=date_from)
    if date_to: queryset = queryset.filter(fecha_solicitud_cita__lte=date_to)
    if status: queryset = queryset.filter(estado_solicitud__icontains=status)
    if procedure: queryset = queryset.filter(tipo_procedimiento__icontains=procedure)
    if eps: queryset = queryset.filter(entidad_aseguradora=eps)
    if barrier: queryset = queryset.filter(barrera=barrier)
    
    if q_search:
        queryset = queryset.filter(
            Q(patient__numero_documento__icontains=q_search) | 
            Q(patient__nombre_1__icontains=q_search) |
            Q(patient__apellido_1__icontains=q_search) |
            Q(observaciones__icontains=q_search) |
            Q(cups__icontains=q_search)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Seguimiento Vidanova"

    headers = [
        'Documento', 'Paciente', 'Edad', 'Género', 'Aseguradora (EPS)', 
        'Procedimiento', 'CUPS', 'Estado', 'F. Solicitud', 'F. Cita', 
        'Días Gestión', 'Barrera', 'Observaciones'
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid") 
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for f in queryset:
        row = [
            f.patient.numero_documento,
            f.patient.nombre_completo,
            f.patient.edad,
            f.patient.genero,
            f.entidad_aseguradora,
            f.tipo_procedimiento,
            f.cups,
            f.estado_solicitud,
            f.fecha_solicitud_cita,
            f.fecha_cita,
            f.dias_diff,
            f.barrera,
            f.observaciones,
        ]
        ws.append(row)

    ws.auto_filter.ref = ws.dimensions
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Gerencial_Vidanova.xlsx"'
    wb.save(response)
    return response

# --- 4. ANÁLISIS GERENCIAL (LA QUE FALTABA) ---
@login_required
def analisis_institucional(request):
    metrics = compute_institutional_metrics_db()
    total_rows = FollowUp.objects.count()
    context = {
        **metrics,
        'rows': total_rows
    }
    return render(request, 'analisis_institucional.html', context)

# --- 5. CRUD ---
@login_required
def followup_detail(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    return render(request, 'followup_detail.html', {'followup': followup})

@login_required
def editar_followup(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    if request.method == 'POST':
        form = FollowUpForm(request.POST, instance=followup)
        if form.is_valid():
            nueva_nota = form.cleaned_data.get('nueva_observacion')
            if nueva_nota:
                usuario = request.user.username.title()
                ahora = timezone.now().strftime("%d/%m/%Y %H:%M")
                entrada_bitacora = f"[{ahora} - {usuario}]: {nueva_nota}"
                historial_previo = followup.observaciones or ""
                followup.observaciones = f"{entrada_bitacora}\n{historial_previo}"
            
            form.save()
            messages.success(request, f"Gestión de {followup.patient.nombre_completo} registrada.")
            return redirect('followup_dashboard')
    else:
        form = FollowUpForm(instance=followup)
    
    return render(request, 'followup_form.html', {
        'form': form, 
        'title': 'Gestionar Caso',
        'patient': followup.patient
    })

@login_required
def agregar_followup(request, patient_id):
    from patients.models import Patient
    patient = get_object_or_404(Patient, pk=patient_id)
    if request.method == 'POST':
        form = FollowUpForm(request.POST)
        if form.is_valid():
            nuevo = form.save(commit=False)
            nuevo.patient = patient
            nuevo.save()
            messages.success(request, "Nuevo seguimiento creado.")
            return redirect('followup_dashboard')
    else:
        form = FollowUpForm(initial={'patient': patient})

    return render(request, 'followup_form.html', {
        'form': form, 
        'title': 'Nuevo Seguimiento', 
        'patient': patient
    })

@login_required
def eliminar_followup(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    if request.method == 'POST':
        followup.delete()
        messages.success(request, "Registro eliminado.")
        return redirect('followup_dashboard')
    return render(request, 'followup_confirm_delete.html', {'followup': followup})

# --- 6. ACCIONES MASIVAS ---

@login_required
def actualizacion_masiva(request):
    """
    Procesa acciones masivas: Estado, Fechas, Clasificación, Barreras y Bitácora.
    """
    if request.method == 'POST':
        ids_str = request.POST.get('selected_ids', '')
        
        # 1. Captura de todos los campos del Modal
        nuevo_estado = request.POST.get('bulk_status')
        nueva_fecha_cita = request.POST.get('bulk_date_cita')
        nueva_fecha_solicitud = request.POST.get('bulk_date_solicitud') # <--- NUEVO
        
        nuevo_tipo_paciente = request.POST.get('bulk_patient_type')
        nuevo_prestador = request.POST.get('bulk_provider')
        nueva_barrera = request.POST.get('bulk_barrier') # <--- NUEVO
        
        nueva_nota = request.POST.get('bulk_observation')
        
        if not ids_str:
            messages.warning(request, "⚠️ No se seleccionaron registros.")
            return redirect('followup_dashboard')

        ids_list = ids_str.split(',')
        registros = FollowUp.objects.filter(id__in=ids_list)
        count = registros.count()
        updated_objs = []

        usuario = request.user.username.title()
        ahora = timezone.now().strftime("%d/%m/%Y %H:%M")

        # 2. Aplicar cambios en memoria
        for r in registros:
            cambios = False

            if nuevo_estado:
                r.estado_solicitud = nuevo_estado
                cambios = True
            
            if nueva_fecha_cita:
                r.fecha_cita = nueva_fecha_cita
                cambios = True

            if nueva_fecha_solicitud:
                r.fecha_solicitud_cita = nueva_fecha_solicitud
                cambios = True
                
            if nuevo_tipo_paciente:
                r.tipo_paciente = nuevo_tipo_paciente
                cambios = True
                
            if nuevo_prestador:
                r.prestador = nuevo_prestador
                cambios = True

            if nueva_barrera:
                r.barrera = nueva_barrera
                cambios = True

            if nueva_nota:
                entrada = f"[{ahora} - {usuario} - MASIVO]: {nueva_nota}"
                historial_previo = r.observaciones or ""
                r.observaciones = f"{entrada}\n{historial_previo}"
                cambios = True
            
            if cambios:
                updated_objs.append(r)

        # 3. Guardar en BD
        if updated_objs:
            fields = ['observaciones']
            if nuevo_estado: fields.append('estado_solicitud')
            if nueva_fecha_cita: fields.append('fecha_cita')
            if nueva_fecha_solicitud: fields.append('fecha_solicitud_cita')
            if nuevo_tipo_paciente: fields.append('tipo_paciente')
            if nuevo_prestador: fields.append('prestador')
            if nueva_barrera: fields.append('barrera')
            
            FollowUp.objects.bulk_update(updated_objs, fields)
            messages.success(request, f"✅ Se actualizaron {count} pacientes correctamente.")
        else:
            messages.info(request, "No se aplicaron cambios.")
            
    return redirect('followup_dashboard')

# --- 7. AUDITORÍA DE DATOS ---
@login_required
def auditoria_calidad(request):
    total_registros = FollowUp.objects.count()
    total_pacientes = FollowUp.objects.values('patient').distinct().count()
    cups_distintos = FollowUp.objects.values('cups').distinct().count()

    posibles_duplicados = FollowUp.objects.values(
        'patient__numero_documento', 'patient__nombre_1', 
        'fecha_solicitud_cita', 'tipo_procedimiento'
    ).annotate(cantidad=Count('id')).filter(cantidad__gt=1).order_by('-cantidad')[:50]

    sin_eps = FollowUp.objects.filter(Q(entidad_aseguradora__isnull=True) | Q(entidad_aseguradora__exact='')).count()
    sin_cups = FollowUp.objects.filter(Q(cups__isnull=True) | Q(cups__exact='')).count()
    fechas_malas = FollowUp.objects.filter(fecha_cita__lt=F('fecha_solicitud_cita')).count()

    context = {
        'total': total_registros,
        'pacientes': total_pacientes,
        'cups_unicos': cups_distintos,
        'duplicados': posibles_duplicados,
        'sin_eps': sin_eps,
        'sin_cups': sin_cups,
        'fechas_malas': fechas_malas
    }
    return render(request, 'auditoria.html', context)

def ver_datos_siisa(request):
    return redirect('followup_dashboard')

# --- 7. MÓDULO DE CALENDARIO ---

@login_required
def calendar_view(request):
    """Renderiza la página del calendario (el contenedor)."""
    return render(request, 'calendar.html')

@login_required
def calendar_api(request):
    """
    API interna que devuelve las citas en formato JSON para FullCalendar.
    """
    # Solo traemos registros que tengan fecha de cita asignada
    citas = FollowUp.objects.filter(fecha_cita__isnull=False).select_related('patient')
    
    eventos = []
    for c in citas:
        # Definir color según procedimiento
        proc = str(c.tipo_procedimiento).upper()
        color = '#6c757d' # Gris (Default)
        
        if 'CONSULTA' in proc: color = '#3b82f6' # Azul
        elif 'QUIMIO' in proc: color = '#8b5cf6' # Morado
        elif 'RADIO' in proc: color = '#f59e0b'  # Naranja
        elif 'CIRUGIA' in proc: color = '#ef4444' # Rojo
        elif 'IMAGEN' in proc: color = '#10b981'  # Verde
        elif 'LABORATORIO' in proc: color = '#0ea5e9' # Azul Cielo

        eventos.append({
            'title': f"{c.patient.nombre_1} {c.patient.apellido_1} - {c.tipo_procedimiento}",
            'start': c.fecha_cita.isoformat(), # Formato YYYY-MM-DD
            'url': f"/seguimiento/detalle/{c.id}/", # Link al hacer clic
            'color': color,
            'description': f"EPS: {c.entidad_aseguradora}"
        })
    
    return JsonResponse(eventos, safe=False)

# --- 8. CENTRO DE ALERTAS ---
@login_required
def centro_alertas(request):
    """
    Muestra alertas con límite de seguridad y paginación.
    """
    # A. INCONSISTENCIAS (Limitamos a 100 para rendimiento y UX)
    inconsistencias_qs = FollowUp.objects.filter(
        fecha_cita__lt=F('fecha_solicitud_cita')
    ).select_related('patient')
    
    total_inconsistencias = inconsistencias_qs.count()
    # Solo traemos los primeros 100 para no explotar la vista
    inconsistencias = inconsistencias_qs[:100] 

    # B. VENCIDOS (Paginación normal)
    fecha_limite = timezone.now().date() - timedelta(days=30)
    estados_pendientes = ['PENDIENTE', 'EN_GESTION', 'POR_GESTIONAR', 'NO_AUTORIZADO']
    
    vencidos_qs = FollowUp.objects.filter(
        estado_solicitud__in=estados_pendientes,
        fecha_solicitud_cita__lt=fecha_limite,
        fecha_cita__isnull=True
    ).select_related('patient').order_by('fecha_solicitud_cita')

    paginator = Paginator(vencidos_qs, 20) # Aumentamos a 20 por página para aprovechar el scroll
    page_number = request.GET.get('page')
    vencidos_page = paginator.get_page(page_number)

    context = {
        'inconsistencias': inconsistencias,
        'total_inconsistencias': total_inconsistencias, # Para mostrar el número real total
        'vencidos': vencidos_page,
        'total_alertas': total_inconsistencias + vencidos_qs.count()
    }
    return render(request, 'alerts_center.html', context)

# --- 9. CONFIGURACIÓN Y MAESTROS ---
@login_required
def configuracion_cups(request):
    """
    Permite a las jefas clasificar los códigos nuevos que el sistema detectó.
    """
    # Si enviaron un formulario para clasificar
    if request.method == 'POST':
        cup_id = request.POST.get('cup_id')
        nuevo_grupo = request.POST.get('new_group')
        
        if cup_id and nuevo_grupo:
            # Actualizamos el maestro
            MasterCUP.objects.filter(id=cup_id).update(grupo=nuevo_grupo)
            messages.success(request, "✅ Código clasificado correctamente.")
            return redirect('configuracion_cups')

    # Listar solo los pendientes
    pendientes = MasterCUP.objects.filter(grupo='PENDIENTE').order_by('codigo')
    
    # Opciones de categorías (Las mismas del modelo)
    categorias = [
        ('CONSULTA', 'Consulta Especializada'),
        ('QUIMIOTERAPIA', 'Quimioterapia'),
        ('RADIOTERAPIA', 'Radioterapia'),
        ('CIRUGIA', 'Cirugía'),
        ('IMAGEN', 'Imagenología'),
        ('LABORATORIO', 'Laboratorio Clínico'),
        ('DOLOR', 'Clínica del Dolor'),
        ('ESTANCIA', 'Estancia / Hospitalización'),
        ('DIAGNOSTICO', 'Procedimiento Diagnóstico'),
        ('COMPLEMENTARIO', 'Servicio Complementario'),
        ('ONCOLOGIA', 'Oncología General'),
        ('OTROS', 'Otros'),
    ]

    return render(request, 'config_cups.html', {
        'pendientes': pendientes,
        'categorias': categorias,
        'total_pendientes': pendientes.count()
    })

# --- 10. UTILIDADES DE SISTEMA (BACKUP) ---
@login_required
def descargar_backup(request):
    """
    Genera y descarga una copia de seguridad de la base de datos SQLite.
    Solo para superusuarios.
    """
    if not request.user.is_superuser:
        messages.error(request, "No tienes permisos para realizar copias de seguridad.")
        return redirect('followup_dashboard')

    # Ruta del archivo original
    db_path = settings.DATABASES['default']['NAME']
    
    # Nombre del archivo para descargar
    timestamp = timezone.now().strftime("%Y-%m-%d_%H%M")
    filename = f"backup_vidanova_{timestamp}.sqlite3"

    # Abrimos el archivo en modo binario y lo enviamos
    try:
        response = FileResponse(open(db_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        messages.error(request, f"Error generando backup: {str(e)}")
        return redirect('followup_dashboard')